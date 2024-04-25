import os
import zipfile
from datetime import date, timedelta
import openpyxl
from django.contrib.auth.views import LoginView
from django.core.exceptions import ObjectDoesNotExist
from django.http import JsonResponse, HttpResponse
from django.shortcuts import redirect, render
from django.urls import reverse_lazy, reverse
from django.views import View
from django.views.generic import CreateView, ListView, DeleteView, UpdateView, DetailView
from drag_and_drop_app.models import Image
from main.all_competitors_list import create_excel_file
from main.competitor_list import create_xls
from main.forms import CompetitionForm
from main.models import Competition
from roles.models import Competitor, Conductor, AdditionalScore

from users.forms import UserLoginForm


class RulesView(View):
    def get(self, request):
        return render(request, 'main/rules.html')


# Create your views here.
class CompetitionList(ListView):
    """Показываем все соревнования"""
    model = Competition
    template_name = 'main/index.html'
    context_object_name = 'object_competition'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        today = date.today()
        past_limit = today - timedelta(days=7)  # предел для прошедших соревнований (не более 7 дней назад)
        context['past_competitions'] = Competition.objects.filter(date_competition__lt=today,
                                                                  date_competition__gte=past_limit)  # Прошедшие соревнования
        future_competitions = Competition.objects.filter(date_competition__gte=today)  # Будущие соревнования
        context['pre_date'] = Competition.objects.filter(date_competition=None)  # если нет конкретной даты
        for competition in future_competitions:
            competition.status = competition.registration_status()
            context['status'] = competition.status

        context['future_competitions'] = future_competitions

        context['competitors'] = Competitor.objects.all()
        # Добавляем информацию о текущем пользователе в контекст
        context['current_user'] = self.request.user

        return context

    def get_queryset(self):
        return super().get_queryset()

    def post(self, request, *args, **kwargs):
        login_view = LoginView.as_view(template_name='users/login.html')
        return login_view(request, *args, **kwargs)


class CompetitionCreate(CreateView):
    """Дженерик создания соревнований"""
    model = Competition
    form_class = CompetitionForm

    def form_valid(self, form):
        self.object = form.save()
        self.object.owner = self.request.user
        self.object.save()

        return super().form_valid(form)

    def get_success_url(self):
        return reverse('detail_com', kwargs={'pk': self.object.pk})


class CompetitionDelete(DeleteView):
    """Удаление Соревнования"""
    model = Competition
    success_url = reverse_lazy('index')


class CompetitionUpdate(UpdateView):
    model = Competition
    form_class = CompetitionForm

    def get(self, request, *args, **kwargs):
        # Получаем форму для обновления объекта
        self.object = self.get_object()
        form = self.get_form()

        return self.render_to_response(self.get_context_data(form=form))

    def get_success_url(self):
        return reverse('detail_com', kwargs={'pk': self.object.pk})


class CompetitionDetail(DetailView):
    model = Competition
    context_object_name = 'object_competition'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        competition = self.get_object()

        # Фильтрация изображений по соревнованию и классу участия
        images = Image.objects.filter(competition=competition)
        images_by_class = {}
        image_classes = ['ro_dety', 'ro_shenki', 'ro_debut', 'ro_veterany', 'ro_1', 'ro_2', 'ro_3', 'ro_4']
        for image_class in image_classes:
            filtered_images = images.filter(image_class_com=image_class)
            images_by_class[image_class] = filtered_images

        context['images_by_class'] = images_by_class
        context['status'] = competition.registration_status()
        context['competitors'] = competition.competitor_set.all()

        context['competitors_ro_dety'] = competition.competitor_set.filter(class_comp='ro_dety')
        context['competitors_ro_shenki'] = competition.competitor_set.filter(class_comp='ro_shenki')
        context['competitors_ro_debut'] = competition.competitor_set.filter(class_comp='ro_debut')
        context['competitors_ro_veterany'] = competition.competitor_set.filter(class_comp='ro_veterany')
        context['competitors_ro_1'] = competition.competitor_set.filter(class_comp='ro_1')
        context['competitors_ro_2'] = competition.competitor_set.filter(class_comp='ro_2')
        context['competitors_ro_3'] = competition.competitor_set.filter(class_comp='ro_3')
        context['competitors_ro_4'] = competition.competitor_set.filter(class_comp='ro_4')

        # Передаем pk в контекст
        context['pk'] = self.kwargs['pk']

        # Проверяем, является ли текущий пользователь участником соревнования
        if self.request.user.is_authenticated:
            user = self.request.user
            context['is_participant'] = competition.competitor_set.filter(user=user).exists()
        else:
            context['is_participant'] = False

    # Проверяем, является ли текущий пользователь создателем соревнования
        is_owner = self.request.user == competition.owner
        context['is_owner'] = is_owner
        # Проводник юзера для проверки
        if self.request.user.is_authenticated:
            try:
                conductor = Conductor.objects.get(user=self.request.user)
                context['conductor'] = conductor
            except Conductor.DoesNotExist:
                context['conductor'] = False

        # Получаем pk текущего пользователя
        current_user_pk = self.request.user.pk
        context['current_user_pk'] = current_user_pk

        return context

    def post(self, request, *args, **kwargs):
        competition = self.get_object()

        class_type = request.POST.get('class_type')
        if class_type == 'all_class':
            competitors = competition.competitor_set.all()

            # Открываем файл и отправляем его как ответ
            output_excel_path = 'home/a/allagro/site1/media/output.xlsx'
            create_excel_file(competitors, output_excel_path)
            with open(output_excel_path, 'rb') as excel_file:
                response = HttpResponse(excel_file.read(),
                                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=output.xlsx'

        else:
            competitors = competition.competitor_set.filter(class_comp=class_type)
            images = Image.objects.filter(competition=competition, image_class_com=class_type)
            icon_list = list(images.values_list('image_list', flat=True))

            # Создаем zip-архив и добавляем в него все файлы Excel
            points_list_zip_path = 'home/a/allagro/site1/media/points_list.zip'
            with zipfile.ZipFile(points_list_zip_path, 'w') as zip_file:
                for competitor in competitors:
                    # print(competition.judge_competition, class_type, competitor.user.first_name,
                    #       competitor.selected_dog.dog_name, icon_list)

                    create_xls(competition.judge_competition, class_type, competitor.user.first_name,
                               competitor.selected_dog.dog_name, icon_list)
                    excel_file_path = 'home/a/allagro/site1/media/points_list.xlsx'
                    file_name = f"points_list_{competitor.user.pk}.xlsx"
                    zip_file.write(excel_file_path, file_name)

            # Открываем zip-архив для чтения в бинарном режиме
            with open(points_list_zip_path, 'rb') as zip_file:
                response = HttpResponse(zip_file, content_type='application/zip')
                response['Content-Disposition'] = 'attachment; filename=points_list.zip'

        return response


class CompetitionResult(View):
    """Результаты соревнования"""

    def get(self, request, *args, **kwargs):
        # Получаем id соревнования из URL
        competition_id = kwargs.get('pk')
        competition = Competition.objects.get(pk=competition_id)
        # Фильтруем участников по соревнованию
        competitors = Competitor.objects.filter(competition_id=competition_id).order_by('-points')

        competitors_ro_dety = competitors.filter(class_comp='ro_dety').order_by('-points')
        # присваиваю каждому участнику в группе дополнительный параметр
        for competitor in competitors_ro_dety:
            try:
                additional_scores = AdditionalScore.objects.get(competitor_id=competitor.pk)
            except ObjectDoesNotExist:
                additional_scores = None

            competitor.additional_scores = additional_scores
            competitor.save()
        competitors_ro_shenki = competitors.filter(class_comp='ro_shenki').order_by('-points')
        # присваиваю каждому участнику в группе дополнительный параметр
        for competitor in competitors_ro_shenki:
            try:
                additional_scores = AdditionalScore.objects.get(competitor_id=competitor.pk)
            except ObjectDoesNotExist:
                additional_scores = None

            competitor.additional_scores = additional_scores
            competitor.save()  # Сохранить изменения
        competitors_ro_debut = competitors.filter(class_comp='ro_debut').order_by('-points')
        # присваиваю каждому участнику в группе дополнительный параметр
        for competitor in competitors_ro_debut:
            try:
                additional_scores = AdditionalScore.objects.get(competitor_id=competitor.pk)
            except ObjectDoesNotExist:
                additional_scores = None

            competitor.additional_scores = additional_scores
            competitor.save()
        competitors_ro_veterany = competitors.filter(class_comp='ro_veterany').order_by('-points')
        # присваиваю каждому участнику в группе дополнительный параметр
        for competitor in competitors_ro_veterany:
            try:
                additional_scores = AdditionalScore.objects.get(competitor_id=competitor.pk)
            except ObjectDoesNotExist:
                additional_scores = None

            competitor.additional_scores = additional_scores
            competitor.save()
        competitors_ro_1 = competitors.filter(class_comp='ro_1').order_by('-points')
        # присваиваю каждому участнику в группе дополнительный параметр
        for competitor in competitors_ro_1:
            try:
                additional_scores = AdditionalScore.objects.get(competitor_id=competitor.pk)
            except ObjectDoesNotExist:
                additional_scores = None

            competitor.additional_scores = additional_scores
            competitor.save()
        competitors_ro_2 = competitors.filter(class_comp='ro_2').order_by('-points')
        # присваиваю каждому участнику в группе дополнительный параметр
        for competitor in competitors_ro_2:
            try:
                additional_scores = AdditionalScore.objects.get(competitor_id=competitor.pk)
            except ObjectDoesNotExist:
                additional_scores = None

            competitor.additional_scores = additional_scores
            competitor.save()
        competitors_ro_3 = competitors.filter(class_comp='ro_3').order_by('-points')
        # присваиваю каждому участнику в группе дополнительный параметр
        for competitor in competitors_ro_3:
            try:
                additional_scores = AdditionalScore.objects.get(competitor_id=competitor.pk)
            except ObjectDoesNotExist:
                additional_scores = None

            competitor.additional_scores = additional_scores
            competitor.save()
        competitors_ro_4 = competitors.filter(class_comp='ro_4').order_by('-points')
        # присваиваю каждому участнику в группе дополнительный параметр
        for competitor in competitors_ro_4:
            try:
                additional_scores = AdditionalScore.objects.get(competitor_id=competitor.pk)
            except ObjectDoesNotExist:
                additional_scores = None

            competitor.additional_scores = additional_scores
            competitor.save()

        context = {
            'competition': competition,
            'competitors': competitors,

            'competitors_ro_dety': competitors_ro_dety,
            'competitors_ro_shenki': competitors_ro_shenki,
            'competitors_ro_debut':  competitors_ro_debut,
            'competitors_ro_veterany': competitors_ro_veterany,
            'competitors_ro_1': competitors_ro_1,
            'competitors_ro_2': competitors_ro_2,
            'competitors_ro_3': competitors_ro_3,
            'competitors_ro_4': competitors_ro_4,
            # Здесь вы можете добавить дополнительные данные в контекст, если это необходимо
        }

        # Если запрос содержит параметр 'download'
        if request.GET.get('download') == 'excel':
            # Создаем новый Excel файл
            workbook = openpyxl.Workbook()
            # Выбираем активный лист
            worksheet = workbook.active

            # Формируем заголовки таблицы
            headers = ['Место', 'Имя участника', 'Баллы', 'Оценка', 'Время']
            for col_num, header in enumerate(headers, 1):
                # Записываем заголовки в ячейки
                cell = worksheet.cell(row=1, column=col_num)
                cell.value = header

            # Заполняем таблицу данными из контекста
            row_num = 2  # Начинаем заполнять таблицу со второй строки
            classes = [
                ('РО-дети', competitors_ro_dety),
                ('РО-щенки', competitors_ro_shenki),
                ('РО-дебют', competitors_ro_debut),
                ('РО-ветераны', competitors_ro_veterany),
                ('РО-1', competitors_ro_1),
                ('РО-2', competitors_ro_2),
                ('РО-3', competitors_ro_3),
                ('РО-4', competitors_ro_4),
            ]
            for class_name, competitors_class in classes:
                # Записываем название класса в первую ячейку строки
                cell = worksheet.cell(row=row_num, column=1)
                cell.value = class_name
                cell.font = openpyxl.styles.Font(bold=True)

                # Форматируем ячейку с названием класса
                worksheet.merge_cells(
                    start_row=row_num, start_column=1,
                    end_row=row_num, end_column=len(headers)
                )
                row_num += 1
                place = 1

                for competitor in competitors_class:
                    row = [place, competitor.user.first_name, competitor.points, competitor.grade_competitor,
                           '{}:{}'.format(competitor.min_time_competitor, competitor.sec_time_competitor)]
                    for col_num, value in enumerate(row, 1):
                        cell = worksheet.cell(row=row_num, column=col_num)
                        cell.value = value
                    row_num += 1
                    place += 1

                # Автонастройка ширины столбцов
            for column_cells in worksheet.columns:
                max_length = 0
                column = column_cells[0].column_letter
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = max_length + 2
                worksheet.column_dimensions[column].width = adjusted_width

                # Возвращаем Excel файл пользователю
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=competition_results.xlsx'
            workbook.save(response)
            return response

        # Если запрос не содержит параметра 'download', возвращаем шаблон с таблицей
        return render(request, 'main/competition_result.html', context)


class CompetitionArchiveView(ListView):
    """Показываем архив прошедших соревнований"""
    model = Competition
    template_name = 'main/competition_archive.html'
    context_object_name = 'past_competitions'

    def get_queryset(self):
        today = date.today()
        return Competition.objects.filter(date_competition__lt=today)