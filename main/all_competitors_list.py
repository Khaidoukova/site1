import os

import openpyxl
from openpyxl.styles import PatternFill


def create_excel_file(competitors, file_path):
    if os.path.exists(file_path):
        # Если файл уже существует, удаляем его, чтобы перезаписать новыми данными
        os.remove(file_path)

    wb = openpyxl.Workbook()
    ws = wb.active

    # Названия столбцов
    headers = ["№", "Проводник", "Собака", "Порода", "Класс участия", "Город", "Телефон", "Email"]

    # Записываем заголовки и определяем ширину колонок
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=idx, value=header)
        column = ws.column_dimensions[openpyxl.utils.get_column_letter(idx)]
        column_width = len(header) + 2  # Добавляем немного дополнительного пространства
        column.width = column_width

    # Устанавливаем цвет для первой строки
    for cell in ws["1:1"]:
        cell.fill = PatternFill(start_color="6AB8EF", end_color="6AB8EF", fill_type="solid")

    # Записываем данные участников
    for idx, competitor in enumerate(competitors, start=2):
        ws.cell(row=idx, column=1, value=idx - 1)
        ws.cell(row=idx, column=2, value=competitor.user.first_name)
        ws.cell(row=idx, column=3, value=competitor.selected_dog.dog_name)
        ws.cell(row=idx, column=4, value=competitor.selected_dog.breed_dog)  # Порода

        # Проверяем наличие класса участия
        if competitor.class_comp:
            ws.cell(row=idx, column=5, value=competitor.class_comp)
        else:
            ws.cell(row=idx, column=5, value=None)

        ws.cell(row=idx, column=6, value=competitor.user.user_town)
        ws.cell(row=idx, column=7, value=competitor.user.phone)
        ws.cell(row=idx, column=8, value=competitor.user.email)

        # Определяем ширину колонки по максимальной длине данных
        for col_num, value in enumerate([idx - 1, competitor.user.first_name, competitor.selected_dog.dog_name,
                                         competitor.selected_dog.breed_dog,
                                         competitor.class_comp, competitor.user.user_town, competitor.user.phone,
                                         competitor.user.email], start=1):
            column = ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)]
            column_width = max(column.width, len(str(value)) + 2)  # Добавляем немного дополнительного пространства
            column.width = column_width

            # Устанавливаем цвет для столбца "Класс участия"
        for cell in ws["E:E"]:
            cell.fill = PatternFill(start_color="00FFFF00", end_color="00FFFF00", fill_type="solid")

    wb.save(file_path)
