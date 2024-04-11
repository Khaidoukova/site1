import openpyxl


def create_xls(judge, class_type, name, dog_name, icon_list):
    """Создаем файл эксель с нужными данными"""
    # Путь к существующему файлу Excel
    excel_file_path = '/home/a/allagro/site1/media/points_list.xlsx'

    # Загрузка существующего файла Excel
    wb = openpyxl.load_workbook(excel_file_path)

    # Выбираем активный лист
    ws = wb.active

    # Записываем данные в конкретные ячейки
    ws['B2'] = class_type  # класс соревнования
    ws['B3'] = judge  # Судья
    ws['B4'] = name  # Проводник
    ws['B5'] = dog_name  # собака
    # Очищаем ячейки B8 - B27
    for row_num in range(8, 28):
        ws[f'B{row_num}'] = ''
    numb = 8
    result = []
    for icon2 in icon_list:
        # Разбиваем строку по запятой и пробелу, но только если перед запятой нет символа '№'
        items = [item.strip() for item in icon2.split(',')]
        # Оставляем только элементы, которые начинаются с символа '№'
        result = [item for item in items if item.startswith('№')]

    for i in result:
        ws[f'B{numb}'] = i
        numb += 1

    # Сохраняем книгу Excel
    wb.save(excel_file_path)

    print("Данные успешно дописаны в файл Excel.")

